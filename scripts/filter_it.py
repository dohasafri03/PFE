#!/usr/bin/env python3
"""
Filtre IT : ne garde que les consultations liées aux 6 domaines IT cibles.
Domaines : AI, Data, BI, Dev, Cloud, Cybersecurity

Entrée  : CSV brut multi-sources (777+ lignes)
Sorties : CSV filtré IT + Excel formaté (onglets par domaine + stats)
"""
import csv
import re
import sys
from collections import Counter
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None  # type: ignore
    Font = PatternFill = Alignment = Border = Side = None  # type: ignore
    get_column_letter = None  # type: ignore

# ── Domaines et mots-clés ──────────────────────────────────────────────────
DOMAINS = {
    "AI": [
        "intelligence artificielle", "machine learning", "deep learning",
        "artificial intelligence", "chatbot", "nlp", "neural",
        "apprentissage automatique", "vision par ordinateur", "modèle prédictif",
        "réseau de neurones", "traitement du langage", "computer vision",
    ],
    "Data": [
        "data", "données", "donnees", "big data", "base de données",
        "database", "sql", "hadoop", "spark", "datawarehouse", "data warehouse",
        "entrepôt de données", "lac de données", "data lake", "analytics",
        "traitement de données", "mongodb", "nosql", "oracle", "postgresql",
        "migration de données", "gestion des données",
    ],
    "BI": [
        "business intelligence", "tableau de bord", "dashboard", "reporting",
        "décisionnel", "power bi", "qlik", "tableau", "dataviz",
        "visualisation de données", "indicateurs de performance", "kpi",
        "aide à la décision", "pilotage",
    ],
    "Dev": [
        "développement", "developpement", "development", "logiciel", "software",
        "application mobile", "application web", "site web", "plateforme numérique",
        "platform", "erp", "crm", "progiciel", "api", "microservice", "devops",
        "intégration", "integration", "système d'information",
        "systeme d'information", "conception", "programmation", "codage",
        "maintenance informatique", "infogérance", "infogerance",
        "transformation numérique", "transformation digitale", "digitalisation",
        "portail web", "portail numérique", "refonte",
        "infrastructure informatique", "réseau informatique",
        "audit informatique", "hébergement web", "hébergement cloud",
    ],
    "Cloud": [
        "cloud", "saas", "paas", "iaas", "virtualisation", "datacenter",
        "data center", "vmware", "azure", "aws", "infrastructure cloud",
        "serveur virtuel", "conteneur", "docker", "kubernetes",
        "sauvegarde cloud", "stockage cloud", "migration cloud",
    ],
    "Cybersecurity": [
        "cybersécurité", "cybersecurite", "sécurité informatique",
        "securite informatique", "firewall", "antivirus", "pare-feu",
        "intrusion", "pentest", "soc", "siem", "cryptage", "chiffrement",
        "audit sécurité", "protection des données", "ransomware",
        "solution de sécurité",
    ],
}

# Short keywords needing word-boundary checks
SHORT_KW = {
    "sql", "nlp", "erp", "crm", "api", "kpi", "soc", "etl",
    "data", "saas", "paas", "iaas", "aws", "siem",
}

DOMAIN_COLORS = {
    "AI":             ("7B2D8E", "F3E5F5"),
    "Data":           ("1565C0", "E3F2FD"),
    "BI":             ("2E7D32", "E8F5E9"),
    "Dev":            ("E65100", "FFF3E0"),
    "Cloud":          ("00838F", "E0F7FA"),
    "Cybersecurity":  ("C62828", "FFEBEE"),
}


def classify(text: str) -> list:
    """Classify text into matching IT domains."""
    if not text:
        return []
    text_lower = text.lower()
    matched = []
    for domain, keywords in DOMAINS.items():
        for kw in keywords:
            kw_clean = kw.strip().lower()
            if kw_clean in SHORT_KW:
                if re.search(r'\b' + re.escape(kw_clean) + r'\b', text_lower):
                    matched.append(domain)
                    break
            elif kw_clean in text_lower:
                matched.append(domain)
                break
    return sorted(set(matched))


def _clean_source(source_raw: str) -> str:
    """Nettoie la colonne source (supprime les doublons répétés)."""
    if not source_raw:
        return ""
    s = source_raw.lower()
    if "marchespublics" in s:
        return "marchespublics.gov.ma"
    if "lesoffres" in s:
        return "lesoffres.ma"
    if "moroccotenders" in s:
        return "moroccotenders.com"
    return source_raw[:50]


def _gen_desc_technique(row: dict) -> str:
    """Génère une description technique à partir des champs disponibles."""
    parts = []
    nature = row.get("nature_prestation", "").strip()
    categorie = row.get("categorie", "").strip()
    articles = row.get("articles", "").strip()

    if nature:
        parts.append(f"Nature: {nature}")
    if categorie:
        parts.append(f"Catégorie: {categorie}")

    if articles:
        items = [a.strip() for a in articles.split("|") if a.strip()]
        if items:
            lots = []
            for item in items[:10]:
                cleaned = re.sub(r'^#\d+\s*', '', item).strip()
                if cleaned:
                    lots.append(f"  - {cleaned}")
            if lots:
                parts.append(f"Lots/Articles ({len(items)}):")
                parts.extend(lots)
                if len(items) > 10:
                    parts.append(f"  ... et {len(items) - 10} autres")

    fichiers = row.get("fichiers_joints", "").strip()
    if fichiers and fichiers not in ("", "Télécharger"):
        parts.append(f"Documents: {fichiers}")

    if not parts:
        objet = row.get("objet", "").strip()
        parts.append(objet if objet else "Non spécifié")

    return "\n".join(parts)


def _gen_desc_fonctionnelle(row: dict) -> str:
    """Génère une description fonctionnelle résumée."""
    parts = []
    objet = row.get("objet", "").strip()
    acheteur = row.get("acheteur", "").strip()
    lieu = row.get("lieu_execution", "").strip().strip("- ")
    budget = row.get("budget_estime", "").strip()
    date_pub = row.get("date_publication", "").strip()
    date_lim = row.get("date_limite", "").strip()
    nb_art = row.get("nb_articles", "").strip()

    if objet:
        parts.append(f"Objet: {objet}")
    if acheteur:
        parts.append(f"Acheteur: {acheteur}")
    if lieu:
        parts.append(f"Lieu: {lieu}")
    if budget:
        parts.append(f"Budget estimé: {budget}")
    if date_pub and date_lim:
        parts.append(f"Période: {date_pub} -> {date_lim}")
    elif date_lim:
        parts.append(f"Date limite: {date_lim}")
    if nb_art and nb_art != "0":
        parts.append(f"Nombre de lots: {nb_art}")

    return "\n".join(parts) if parts else "Non spécifié"


def filter_and_export(csv_input=None):
    """
    Filtre le CSV pour ne garder que les consultations IT.
    Exporte : CSV filtré + Excel formaté.
    Retourne dict avec stats.
    """
    if csv_input:
        csv_path = Path(csv_input)
    else:
        data_dir = Path("data")
        csvs = sorted(data_dir.glob("appels_offres_*.csv"), reverse=True)
        if not csvs:
            print("[ERR] Aucun CSV trouvé dans data/")
            return None
        csv_path = csvs[0]

    # ── Lecture CSV ────────────────────────────────────────────────────────
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        all_rows = list(reader)

    print(f"  CSV lu: {len(all_rows)} consultations brutes")

    # ── Classification + Filtrage IT ────────────────────────────────────────
    it_rows = []
    domain_rows = {d: [] for d in DOMAINS}

    for row in all_rows:
        row["source"] = _clean_source(row.get("source", ""))

        text = " ".join([
            row.get("objet", ""),
            row.get("articles", ""),
            row.get("nature_prestation", ""),
            row.get("categorie", ""),
        ])
        cats = classify(text)

        if not cats:
            continue  # ← ON NE GARDE QUE LES CONSULTATIONS IT

        row["domaines_it"] = ", ".join(cats)
        row["description_technique"] = _gen_desc_technique(row).replace("\n", " | ")
        row["description_fonctionnelle"] = _gen_desc_fonctionnelle(row).replace("\n", " | ")

        it_rows.append(row)
        for c in cats:
            domain_rows[c].append(row)

    print(f"  Consultations IT retenues: {len(it_rows)} / {len(all_rows)}")
    for d in DOMAINS:
        cnt = len(domain_rows[d])
        if cnt:
            print(f"    {d:15s}: {cnt:>4d}")

    # ── Export CSV filtré (écrase le source + crée _IT) ───────────────────
    fieldnames = [
        "reference", "objet", "acheteur", "categorie", "nature_prestation",
        "date_publication", "date_limite", "lieu_execution", "budget_estime",
        "description_technique", "description_fonctionnelle",
        "domaines_it", "source", "url",
    ]

    # 1) Écraser le CSV source avec uniquement les lignes IT
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";",
                                extrasaction="ignore")
        writer.writeheader()
        writer.writerows(it_rows)
    print(f"  [CSV]   {csv_path} (ecrase -> {len(it_rows)} lignes IT)")

    # 2) Copie _IT pour référence
    csv_filtered = csv_path.parent / csv_path.name.replace(".csv", "_IT.csv")
    with open(csv_filtered, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";",
                                extrasaction="ignore")
        writer.writeheader()
        writer.writerows(it_rows)
    print(f"  [CSV]   {csv_filtered} ({len(it_rows)} lignes IT)")

    # ── Export Excel (optional) ───────────────────────────────────────────
    # If openpyxl isn't installed, we still return the filtered CSV (n8n should not fail).
    if openpyxl is None:
        return {
            "csv_source": str(csv_path),
            "csv_filtered": str(csv_filtered),
            "excel": None,
            "total_brutes": len(all_rows),
            "total_it": len(it_rows),
            "par_domaine": {d: len(domain_rows[d]) for d in DOMAINS},
            "message": "Excel export skipped (openpyxl not installed).",
        }

    xlsx_path = csv_path.parent / "consultations_IT.xlsx"

    display_cols = [
        ("reference",                "RÉFÉRENCE",                18),
        ("objet",                    "OBJET",                    55),
        ("acheteur",                 "ACHETEUR",                 35),
        ("categorie",                "CATÉGORIE",                14),
        ("date_publication",         "DATE PUBLICATION",         15),
        ("date_limite",              "DATE LIMITE",              15),
        ("lieu_execution",           "LIEU D'EXÉCUTION",         20),
        ("budget_estime",            "BUDGET ESTIMÉ",            18),
        ("description_technique",    "DESCRIPTION TECHNIQUE",    50),
        ("description_fonctionnelle","DESCRIPTION FONCTIONNELLE",50),
        ("domaines_it",              "DOMAINES IT",              22),
        ("source",                   "SOURCE",                   22),
        ("url",                      "URL",                      45),
    ]

    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    wrap_align = Alignment(vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()

    def write_sheet(ws, data_rows, header_color="2F5496", row_fill_color=None):
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        for col_idx, (field, label, width) in enumerate(display_cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        r_fill = PatternFill(start_color=row_fill_color, end_color=row_fill_color, fill_type="solid") if row_fill_color else None
        for row_idx, row in enumerate(data_rows, 2):
            for col_idx, (field, _, _) in enumerate(display_cols, 1):
                val = row.get(field, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = wrap_align
                cell.border = thin_border
                if r_fill:
                    cell.fill = r_fill

        ws.freeze_panes = "A2"
        if data_rows:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(display_cols))}{len(data_rows)+1}"

    # ── Onglet principal — Toutes IT ──────────────────────────────────────
    ws_all = wb.active
    ws_all.title = f"Toutes IT ({len(it_rows)})"
    write_sheet(ws_all, it_rows)

    # Colorer par domaine principal
    for row_idx, row in enumerate(it_rows, 2):
        domain_str = row.get("domaines_it", "")
        if domain_str:
            primary = domain_str.split(",")[0].strip()
            if primary in DOMAIN_COLORS:
                _, bg = DOMAIN_COLORS[primary]
                fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                for col_idx in range(1, len(display_cols) + 1):
                    ws_all.cell(row=row_idx, column=col_idx).fill = fill

    # ── Onglets par domaine IT ─────────────────────────────────────────────
    for domain in DOMAINS:
        dr = domain_rows[domain]
        if not dr:
            continue
        ws = wb.create_sheet(title=f"{domain} ({len(dr)})")
        fg, bg = DOMAIN_COLORS[domain]
        write_sheet(ws, dr, header_color=fg, row_fill_color=bg)

    # ── Onglet Statistiques ────────────────────────────────────────────────
    ws_stats = wb.create_sheet(title="Statistiques")
    ws_stats.column_dimensions["A"].width = 32
    ws_stats.column_dimensions["B"].width = 15
    ws_stats.column_dimensions["C"].width = 12

    r = 1
    ws_stats.cell(row=r, column=1, value="STATISTIQUES - APPELS D'OFFRES IT").font = Font(bold=True, size=14, color="2F5496")
    r += 2
    ws_stats.cell(row=r, column=1, value="Total consultations scrapées").font = Font(bold=True)
    ws_stats.cell(row=r, column=2, value=len(all_rows))
    r += 1
    ws_stats.cell(row=r, column=1, value="Consultations IT retenues").font = Font(bold=True)
    ws_stats.cell(row=r, column=2, value=len(it_rows))
    r += 1
    ws_stats.cell(row=r, column=1, value="Taux de filtrage IT").font = Font(bold=True)
    ws_stats.cell(row=r, column=2, value=f"{len(it_rows)*100//max(len(all_rows),1)}%")
    r += 2

    ws_stats.cell(row=r, column=1, value="PAR DOMAINE IT").font = Font(bold=True, size=12)
    r += 1
    ws_stats.cell(row=r, column=1, value="Domaine").font = Font(bold=True)
    ws_stats.cell(row=r, column=2, value="Nb").font = Font(bold=True)
    ws_stats.cell(row=r, column=3, value="%").font = Font(bold=True)
    r += 1
    for domain in DOMAINS:
        cnt = len(domain_rows[domain])
        pct = round(cnt / max(len(it_rows), 1) * 100, 1)
        fg, bg = DOMAIN_COLORS[domain]
        ws_stats.cell(row=r, column=1, value=domain)
        ws_stats.cell(row=r, column=1).fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        ws_stats.cell(row=r, column=1).font = Font(bold=True, color=fg)
        ws_stats.cell(row=r, column=2, value=cnt)
        ws_stats.cell(row=r, column=3, value=f"{pct}%")
        r += 1

    r += 1
    ws_stats.cell(row=r, column=1, value="PAR SOURCE").font = Font(bold=True, size=12)
    r += 1
    for s, c in Counter(row.get("source", "autre") for row in it_rows).most_common():
        ws_stats.cell(row=r, column=1, value=s)
        ws_stats.cell(row=r, column=2, value=c)
        r += 1

    r += 1
    ws_stats.cell(row=r, column=1, value="PAR CATÉGORIE").font = Font(bold=True, size=12)
    r += 1
    for cat, c in Counter(row.get("categorie", "N/A") or "N/A" for row in it_rows).most_common():
        ws_stats.cell(row=r, column=1, value=cat)
        ws_stats.cell(row=r, column=2, value=c)
        r += 1

    try:
        wb.save(xlsx_path)
    except PermissionError:
        # File locked (e.g. open in Excel) — save with timestamp suffix
        from datetime import datetime as _dt
        alt = xlsx_path.with_stem(f"consultations_IT_{_dt.now().strftime('%H%M%S')}")
        wb.save(alt)
        xlsx_path = alt
        print(f"  [WARN] Fichier verrouillé, sauvegardé sous: {xlsx_path.name}")
    print(f"  [EXCEL] {xlsx_path}")
    print(f"    -> {len(it_rows)} consultations IT uniquement")
    print(f"    -> {len(DOMAINS)} onglets domaine + Statistiques")

    return {
        "csv_source": str(csv_path),
        "csv_filtered": str(csv_filtered),
        "excel": str(xlsx_path),
        "total_brutes": len(all_rows),
        "total_it": len(it_rows),
        "par_domaine": {d: len(domain_rows[d]) for d in DOMAINS},
    }


def main():
    import argparse
    parser = argparse.ArgumentParser(description='Filtre IT : CSV -> CSV filtré + Excel')
    parser.add_argument('csv', nargs='?', default=None, help='Chemin CSV (auto-détect si omis)')
    args = parser.parse_args()
    result = filter_and_export(args.csv)
    if result:
        print(f"\n  Résultat: {result['total_it']} IT / {result['total_brutes']} totales")


if __name__ == "__main__":
    main()
